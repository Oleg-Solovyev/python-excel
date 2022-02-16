# 2022-01-28  oleg.s.solovyev@gmail.com  new code
# updates Excel file with formulas


# toDo
#   loop over io3 spread sheet
#   get values for C9 column
#   double check that .replace('-', '').lower() works for every case


# libraries 
import openpyxl
import itertools
import string


# constants
INPUT_EXCEL_FILE_PATH  = 'equations.xlsx'
OUTPUT_EXCEL_FILE_PATH = 'results.xlsx'
# https://stackoverflow.com/questions/42176498/repeating-letters-like-excel-columns
EXCEL_COLUMNS_LIST = list(itertools.chain(string.ascii_uppercase, (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))))


# open Excel file
wb = openpyxl.load_workbook(INPUT_EXCEL_FILE_PATH)
ws_io3 = wb['io3']
ws_inv = wb['inv']

origin = 'P-50'
destination = 'MSG-RO'


def get_orig_dest_match(workseet, origin, destination):
    
    orig_cand_set = set()
    dest_cand_set = set()
    for row in workseet.iter_rows(min_row=3, max_row=3, max_col=100):
        for cell in row:
            if cell.value == origin:
                orig_cand_set.add(cell.column)
    
    for row in workseet.iter_rows(min_row=4, max_row=4, max_col=100):
        for cell in row:
            if cell.value == destination:
                dest_cand_set.add(cell.column)
    
    match = list(set(orig_cand_set) & set(dest_cand_set))
    if len(match) == 0:
        raise ValueError('no match in inv sheet found for ' + origin + ' and ' + destination)
    elif len(match) > 1:
        raise ValueError('more then 1 match in inv sheet found for ' + origin + ' and ' + destination)
    else:
        print(origin, 'and', destination, 'yield column', EXCEL_COLUMNS_LIST[match[0]-1], 'in inv sheet')
        return match[0]


def get_blue_row(workseet, origin):
    for row in workseet.iter_rows(min_row=13, max_row=42, max_col=1):
        for cell in row:
            if cell.value == origin or cell.value == origin.replace('-', '').lower():
                print(origin, 'yield row', cell.row, 'in blue range')
                return cell.row
    
    raise ValueError('not match for value', origin, 'in blue range')


def get_yellow_row(workseet, origin):
    for row in workseet.iter_rows(min_row=45, max_row=74, max_col=1):
        for cell in row:
            if cell.value == origin or cell.value == origin.replace('-', '').lower():
                print(origin, 'yield row', cell.row, 'in yellow range')
                return cell.row
    
    raise ValueError('not match for value', origin, 'in yellow range')
                
def get_formula(workseet, origin, destination, row):
    match_col_num = get_orig_dest_match(ws_inv, origin, destination)
    blue_row   = get_blue_row(ws_inv, origin)
    yellow_row = get_yellow_row(ws_inv, origin)
    
    match_col_name = EXCEL_COLUMNS_LIST[match_col_num-1]
    
    formula = '= C' +  str(row) + ' - (' + \
              'inv!' + match_col_name + '5*inv!' + match_col_name +  '9*inv!' + match_col_name + str(blue_row  ) + ' - '\
              'inv!' + match_col_name + '6*inv!' + match_col_name + '10*inv!' + match_col_name + str(yellow_row) + \
              ')/1000'
    return(formula)


# loop over io3 sheet
# loop over region
start_row = 2
end_row   = 2
start_col = 4
end_col   = 8


def fill_in_range(ws_io3, ws_inv, start_row, end_row, start_col, end_col):
    for row in ws_io3.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            print('filling in cell ' + EXCEL_COLUMNS_LIST[cell.column-1] + str(cell.row))
            origin = ws_io3.cell(column=2, row=cell.row).value
            destination = ws_io3.cell(column=cell.column, row=start_row-1).value
            try:
                formula = get_formula(ws_io3, origin, destination, cell.row)
                print('formula:', formula)
                ws_io3.cell(column=cell.column, row=cell.row, value=formula)
            except Exception as error:
                print('ERROR:', str(error))

fill_in_range(ws_io3, ws_inv,  2,  4, 4,  8)
fill_in_range(ws_io3, ws_inv,  9, 12, 4,  8)
fill_in_range(ws_io3, ws_inv, 17, 20, 4,  7)
fill_in_range(ws_io3, ws_inv, 23, 25, 4,  9)
fill_in_range(ws_io3, ws_inv, 28, 29, 4,  7)
fill_in_range(ws_io3, ws_inv, 32, 37, 4, 10)



    

# save results
wb.save(OUTPUT_EXCEL_FILE_PATH)
