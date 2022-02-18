# 2022-02-16  oleg.s.solovyev@gmail.com  code copied from update_excel.py
# Description: updates Excel file with formulas
# To lint the code run: $ pycodestyle update_excel_2.py
# To run the docstring tests: $ python3 -m doctest -v update_excel_2.py
# Warning: openpyxl library does not preserve graphs
# after openning and saving Excel file


# libraries
import openpyxl
import itertools
import string


# constants
INPUT_EXCEL_FILE_PATH = '/home/p/Downloads/io2.xlsx'
OUTPUT_EXCEL_FILE_PATH = '/home/p/Downloads/results.xlsx'
# https://stackoverflow.com/questions/42176498/repeating-letters-like-excel-columns
EXCEL_COLUMNS_LIST = list(itertools.chain(string.ascii_uppercase, (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))))


# open Excel file
wb = openpyxl.load_workbook(INPUT_EXCEL_FILE_PATH)


def get_orig_dest_match(sheet, origin, destination):
    '''Get Excel column in sheet inv that matches both origin and destination
    
    >>> get_orig_dest_match(wb['inv'], 'P-20', 'T-P20')
    P-20 and T-P20 yield column B in inv sheet
    2
    >>> get_orig_dest_match(wb['inv'], 'P-20', 'PLEM-AB')
    Traceback (most recent call last):
      File "<stdin>", line 1, in <module>
      File "<stdin>", line 33, in get_orig_dest_match
    ValueError: no match in inv sheet found for P-20 and PLEM-AB
    >>> get_orig_dest_match(wb['inv'], 'Ponto A', 'Cabiúnas')
    Traceback (most recent call last):
      File "<stdin>", line 1, in <module>
      File "<stdin>", line 33, in get_orig_dest_match
    ValueError: more then 1 match in inv sheet found for Ponto A and Cabiúnas
    '''
    
    orig_cand_set = set()
    dest_cand_set = set()
    for row in sheet.iter_rows(min_row=3, max_row=3, max_col=100):
        for cell in row:
            if cell.value == origin:
                orig_cand_set.add(cell.column)
                
    for row in sheet.iter_rows(min_row=4, max_row=4, max_col=100):
        for cell in row:
            if cell.value == destination:
                dest_cand_set.add(cell.column)
                
    match = list(set(orig_cand_set) & set(dest_cand_set))
    if len(match) == 0:
        raise ValueError('no match in inv sheet found for ' +
                         origin + ' and ' + destination)
    elif len(match) > 1:
        raise ValueError('more then 1 match in inv sheet found for ' +
                         origin + ' and ' + destination)
    else:
        print(origin, 'and', destination, 'yield column',
              EXCEL_COLUMNS_LIST[match[0]-1], 'in inv sheet')
        return match[0]


def clear_origin_destination(sheet, min_row, cell_row, cell_column):
    '''Get the origin and destination value for a cell with formula
    >>> clear_origin_destination(wb['A'], 22, 22, 4)
    ('P-20', 'PLEM-AB')
    >>> clear_origin_destination(wb['A'], 22, 22, 6)
    ('PLEM-AB', 'T-P20')
    '''
    
    value = sheet.cell(row=min_row - 1, column=cell_column).value.split(' > ')
    origin = value[0]
    destination = value[-1]
    if origin == 'UEE':
        origin = sheet.cell(column=2, row=cell_row).value
    
    return origin, destination


def get_blue_or_yellow_row(sheet, color, origin):
    '''Get row number of the origin value in column A of the sheet inv, blue region
    >>> get_blue_or_yellow_row(wb['inv'], 'blue', 'P-20')
    P-20 yield row 13 in blue range
    13
    >>> get_blue_or_yellow_row(wb['inv'], 'blue', 'PLEM-AB')
    Traceback (most recent call last):
      File "<stdin>", line 1, in <module>
      File "<stdin>", line 28, in get_blue_or_yellow_row
    ValueError: ('not match for value', 'PLEM-AB', 'in', 'blue', 'range')
    >>> get_blue_or_yellow_row(wb['inv'], 'yellow', 'P-20')
    P-20 yield row 45 in yellow range
    45
    >>> get_blue_or_yellow_row(wb['inv'], 'red', 'P-20')
    Traceback (most recent call last):
      File "<stdin>", line 1, in <module>
      File "<stdin>", line 21, in get_blue_or_yellow_row
    ValueError: ('only blue or yellow color regions are allowed:', 'red')
    '''
    
    if color == 'blue':
        min_row=13
        max_row=42
    elif color == 'yellow':
        min_row=45
        max_row=74
    else:
        raise ValueError('only blue or yellow color regions are allowed:', color)
    
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == origin or cell.value == origin.replace('-', '').lower():
                print(origin, 'yield row', cell.row, 'in', color, 'range')
                return cell.row
    
    raise ValueError('not match for value', origin, 'in', color, 'range')


def fill_in_range(sheet_name, ws_inv, min_row, max_row, min_col, max_col):
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            try:
                print('filling in sheet ' + sheet_name + ' cell ' + EXCEL_COLUMNS_LIST[cell.column-1] + str(cell.row))
                origin, destination = clear_origin_destination(sheet, min_row, cell.row, cell.column)
                match_col_num = get_orig_dest_match(ws_inv, origin, destination)
                
                blue_row = get_blue_or_yellow_row(ws_inv, 'blue', origin)
                yellow_row = get_blue_or_yellow_row(ws_inv, 'yellow', origin)
                
                match_col_name = EXCEL_COLUMNS_LIST[match_col_num-1]
                formula = '= (' + \
                          'inv!' + match_col_name + '5*inv!' + match_col_name +  '9*inv!' + match_col_name + str(blue_row  ) + ' - '\
                          'inv!' + match_col_name + '6*inv!' + match_col_name + '10*inv!' + match_col_name + str(yellow_row) + \
                          ')/1000'
                print('formula:', formula)
                sheet.cell(row=cell.row, column = cell.column, value=formula)
            except Exception as error:
                print('ERROR:', str(error))


#fill_in_range('A', wb['inv'], 22, 24, 4, 9)
#fill_in_range('B', wb['inv'], 21, 26, 4, 8)
#fill_in_range('C', wb['inv'], 26, 29, 4, 7)
#fill_in_range('D', wb['inv'], 22, 24, 4, 8)
#fill_in_range('E', wb['inv'], 20, 21, 4, 7)
#fill_in_range('F', wb['inv'], 24, 28, 4, 7)


#get_orig_dest_match(wb['A'], 'P-20', 'PLEM-AB')


# save results
wb.save(OUTPUT_EXCEL_FILE_PATH)
